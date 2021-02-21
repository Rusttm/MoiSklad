# Подключаем библиотеки
import os
import configparser
import httplib2
import apiclient.discovery
from oauth2client.service_account import ServiceAccountCredentials
from openpyxl import load_workbook
import pandas as pd
import gspread
import gspread_dataframe as gd
from gspread_dataframe import set_with_dataframe, get_as_dataframe



conf = configparser.ConfigParser()
conf.read('config.ini')

CREDENTIALS_FILE = conf['GOOGLE']['CREDENTIALS_FILE']  # Имя файла с закрытым ключом, вы должны подставить свое
spreadsheetId=conf['GOOGLE']['spreadsheetId']
# Читаем ключи из файла
credentials = ServiceAccountCredentials.from_json_keyfile_name(CREDENTIALS_FILE, ['https://www.googleapis.com/auth/spreadsheets', 'https://www.googleapis.com/auth/drive'])
httpAuth = credentials.authorize(httplib2.Http()) # Авторизуемся в системе
service = apiclient.discovery.build('sheets', 'v4', http = httpAuth) # Выбираем работу с таблицами и 4 версию API
spreadsheet = service.spreadsheets().get(spreadsheetId=spreadsheetId).execute()
sheetList = spreadsheet.get('sheets')
sheetName = sheetList[0]['properties']['title'] #'2021y'

def make_sheet():
    '''Return spread sheet id'''
    spreadsheet = service.spreadsheets().create(body={
        'properties': {'title': 'Первый тестовый документ', 'locale': 'ru_RU'},
        'sheets': [{'properties': {'sheetType': 'GRID',
                                   'sheetId': 0,
                                   'title': 'Лист номер один',
                                   'gridProperties': {'rowCount': 100, 'columnCount': 15}}}]
    }).execute()
    spreadsheetId = spreadsheet['spreadsheetId']  # сохраняем идентификатор файла
    print('https://docs.google.com/spreadsheets/d/' + spreadsheetId)
    return spreadsheetId





#делаем доступ к листу
def make_access(spreadsheetId,user_email=conf['GOOGLE']['user_email']):
    driveService = apiclient.discovery.build('drive', 'v3', http = httpAuth) # Выбираем работу с Google Drive и 3 версию API
    access = driveService.permissions().create(
        fileId = spreadsheetId,
        body = {'type': 'user', 'role': 'writer', 'emailAddress': user_email},  # Открываем доступ на редактирование
        fields = 'id'
    ).execute()

def add_sheet(spreadsheetId):
    # Добавление листа
    results = service.spreadsheets().batchUpdate(
        spreadsheetId=spreadsheetId,
        body=
        {
            "requests": [
                {
                    "addSheet": {
                       "properties": {
                           "title": "Еще один лист",
                           "gridProperties": {
                               "rowCount": 20,
                                "columnCount": 12
                            }
                        }
                    }
                }
            ]
        }).execute()
    return results

def get_sheets(spreadsheetId):
    '''send sheetid and Получаем список листов, их Id и название'''
    spreadsheet = service.spreadsheets().get(spreadsheetId=spreadsheetId).execute()
    sheetList = spreadsheet.get('sheets')
    for sheet in sheetList:
        print(sheet['properties']['sheetId'], sheet['properties']['title'])
    sheetId = sheetList[0]['properties']['sheetId']
    return sheetList


def fill_cells(spreadsheetId, sheetName,range: str,value):
    results = service.spreadsheets().values().batchUpdate(spreadsheetId=spreadsheetId, body={
        "valueInputOption": "USER_ENTERED",
        # Данные воспринимаются, как вводимые пользователем (считается значение формул)
        "data": [
            {"range": str(sheetName+'!'+range),
             "majorDimension": "ROWS",  # Сначала заполнять строки, затем столбцы
             "values":  value # Заполняем массивом [[A1,B1,C1],[A2,B2,C2]]
             }
        ]
    }).execute()
    return results







def read_xls(path='./outinvoices_2021_01_01.xlsx'):
    '''read data from file'''
    wb = load_workbook(path)
    sheet = wb['Sheet']
    sheet.max_row   # Retrieve the maximum amount of rows
    sheet.max_column # Retrieve the maximum amount of columns

    for i in range(1,sheet.max_row):
        cell_coord=str('A%s' %i)
        print(sheet[cell_coord].value)

    return True

#fill_cells(spreadsheetId, sheetName, 'A2:A2', [['Hello, World!']])

def write_df_to_datasheet():
    '''push df to google sheet'''
    # ACCES GOOGLE SHEET
    gc = gspread.service_account(filename=CREDENTIALS_FILE)
    sh = gc.open_by_key(spreadsheetId)
    try:
        worksheet = sh.get_worksheet(0)  # -> 0 - first sheet, 1 - second sheet etc.
        print(worksheet)
    except:
        print('cantt find this sheet in book')

    # APPEND DATA TO SHEET
    book = pd.read_excel('supply_2021-02-13.xlsx', sheet_name='покупки')
    set_with_dataframe(worksheet, book)  # -> THIS EXPORTS YOUR DATAFRAME TO THE GOOGLE SHEET

def append_df_to_datasheet():
    # ACCES GOOGLE SHEET
    gc = gspread.service_account(filename=CREDENTIALS_FILE)
    sh = gc.open_by_key(spreadsheetId)
    try:
        worksheet = sh.get_worksheet(0)  # -> 0 - first sheet, 1 - second sheet etc.
    except:
        print('cantt find this sheet in book')

    columns_xls = ['Номер и дата счета-фактуры продавца',
               'Номер и дата исправления счета-фактуры продавца',
               'Номер и дата корректировочного счета-фактуры продавца',
               'Номер и дата исправления корректировочного счета-фактуры продавца',
               'Номер и дата документа, подтверждающего уплату налога',
               'Дата принятия на учет товаров (работ, услуг), имущественных прав',
               'Наименование продавца',
               'ИНН/КПП продавца',
               'Страна происхождения товара. Номер таможенной декларации',
               'Наиме-нование и код валюты',
               'Всего покупок, включая НДС',
               'сумма НДС',
               'покупки, освобождаемые от налога']
    existing = get_as_dataframe(worksheet) #get sheet to dataframe
    print(existing)
    book = pd.read_excel('supply_2021-02-13.xlsx', sheet_name='покупки') #get data from xlsx to dataframe
    finall_df=existing.append(book, ignore_index = True) #append df from xls to sheet
    print(finall_df)
    set_with_dataframe(worksheet, finall_df)  #write df to google sheet

    return True

append_df_to_datasheet()
#write_df_to_datasheet()
